﻿# -*- coding: utf-8 -*-
"""
完税证明PDF批量提取工具 - 图形界面版
功能：
  1. 使用pdfplumber提取PDF文本和表格
  2. 通过正则表达式从文本中提取主表数据（完税证明id、完税金额）
  3. 从表格中解析明细数据（原凭证号、税种、品目名称、税款所属时期、入退库日期、实缴金额）
  4. 输出Excel（.xlsx）和SQL两种格式
  5. 提供图形界面操作，支持实时日志显示
"""

import os
import re
import hashlib
import sys
import glob
import time
import subprocess
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from datetime import datetime
from queue import Queue
from threading import Thread

# ========== 依赖检查 ==========
def check_dependencies():
    """启动时检查依赖包，缺失则自动安装（打包后的exe跳过）"""
    if getattr(sys, 'frozen', False):
        return
    missing = []
    try:
        import pdfplumber
    except ImportError:
        missing.append('pdfplumber')
    try:
        import openpyxl
    except ImportError:
        if 'openpyxl' not in missing:
            missing.append('openpyxl')

    if missing:
        print(f"正在安装缺失的依赖: {', '.join(missing)}")
        for pkg in missing:
            try:
                subprocess.check_call([sys.executable, '-m', 'pip', 'install', pkg])
                print(f"  ✓ {pkg} 安装成功")
            except subprocess.CalledProcessError as e:
                print(f"  ✗ {pkg} 安装失败: {e}")
                sys.exit(1)
        print("依赖安装完成，重新导入中...\n")

# 执行依赖检查
check_dependencies()

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill


# ========== 核心提取函数（复用merge_extract.py） ==========

def extract_no_from_text(text):
    """从PDF文本中提取No.字段（完税证明id）"""
    match = re.search(r'No\.\s*(\d+)', text)
    if match:
        return match.group(1)
    return None


def extract_amount_tax_from_text(text):
    """从PDF文本中提取金额合计（完税金额），格式：￥2,556.42"""
    match = re.search(r'￥([\d,]+\.?\d*)', text)
    if match:
        return match.group(1).replace(',', '')
    return None


def parse_date_range(date_str):
    """
    解析税款所属时期，格式如：2024-07-01至2024-07-31
    返回 (start_date, end_date)
    """
    match = re.match(r'(\d{4}-\d{2}-\d{2})至(\d{4}-\d{2}-\d{2})', date_str)
    if match:
        return match.group(1), match.group(2)
    match = re.match(r'(\d{4}-\d{2}-\d{2})', date_str)
    if match:
        return match.group(1), match.group(1)
    return None, None


def split_merged_cell(cell_text):
    """拆分pdfplumber提取的合并单元格，按换行符分隔多行数据"""
    if not cell_text:
        return []
    lines = [line.strip() for line in cell_text.split('\n') if line.strip()]
    return lines


def extract_table_details(table_data):
    """从pdfplumber提取的表格原始数据中解析明细行"""
    details = []
    if not table_data or len(table_data) < 3:
        return details

    for row in table_data:
        if not row:
            continue
        first_cell = row[0] if row[0] else ''
        if not re.match(r'.*\d{10,}', first_cell):
            continue

        voucher_numbers = split_merged_cell(first_cell)
        tax_categories = split_merged_cell(row[2] if len(row) > 2 else '')
        items_names = split_merged_cell(row[4] if len(row) > 4 else '')
        date_ranges = split_merged_cell(row[6] if len(row) > 6 else '')
        storage_dates = split_merged_cell(row[8] if len(row) > 8 else '')
        amounts = split_merged_cell(row[9] if len(row) > 9 else '')

        max_len = max(len(voucher_numbers), len(tax_categories), len(items_names),
                      len(date_ranges), len(storage_dates), len(amounts))

        for i in range(max_len):
            date_str = date_ranges[i] if i < len(date_ranges) else ''
            start_date, end_date = parse_date_range(date_str)
            detail = {
                'voucher_number': voucher_numbers[i] if i < len(voucher_numbers) else '',
                'tax_categories': tax_categories[i] if i < len(tax_categories) else '',
                'items_name': items_names[i] if i < len(items_names) else '',
                'start_date': start_date,
                'end_date': end_date,
                'storage_date': storage_dates[i] if i < len(storage_dates) else '',
                'amount': amounts[i] if i < len(amounts) else '',
            }
            details.append(detail)

    return details


def extract_single_pdf(pdf_path):
    """提取单个PDF文件的主表和明细数据"""
    main_record = {
        'invoice_number': '',
        'amount_tax': 0.0,
    }
    detail_records = []
    filename = os.path.basename(pdf_path)

    with pdfplumber.open(pdf_path) as pdf:
        if not pdf.pages:
            return main_record, detail_records, filename

        page = pdf.pages[0]
        text = page.extract_text() or ''
        tables = page.extract_tables()

        invoice_number = extract_no_from_text(text)
        if not invoice_number:
            fn_match = re.search(r'(\d{10,})', filename)
            if fn_match:
                invoice_number = fn_match.group(1)
        main_record['invoice_number'] = invoice_number or ''

        amount_tax_str = extract_amount_tax_from_text(text)
        main_record['amount_tax'] = float(amount_tax_str) if amount_tax_str else 0.0

        if tables:
            for table in tables:
                details = extract_table_details(table)
                detail_records.extend(details)

    return main_record, detail_records, filename


# ========== 输出函数 ==========

def generate_excel(all_records, output_path):
    """生成Excel文件"""
    wb = Workbook()
    ws_main = wb.active
    ws_main.title = '主表'

    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font_white = Font(bold=True, size=11, color='FFFFFF')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    main_headers_cn = ['文件名称', '文件路径', '完税证明ID', '完税金额']
    for col_idx, header in enumerate(main_headers_cn, 1):
        cell = ws_main.cell(row=1, column=col_idx, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    for row_idx, rec in enumerate(all_records, 2):
        main = rec['main']
        ws_main.cell(row=row_idx, column=1, value=rec['filename']).border = thin_border
        ws_main.cell(row=row_idx, column=2, value=rec.get('file_path', '')).border = thin_border
        ws_main.cell(row=row_idx, column=3, value=main['invoice_number']).border = thin_border
        amount_cell = ws_main.cell(row=row_idx, column=4, value=main['amount_tax'])
        amount_cell.number_format = '#,##0.00'
        amount_cell.border = thin_border

    ws_main.column_dimensions['A'].width = 20
    ws_main.column_dimensions['B'].width = 40
    ws_main.column_dimensions['C'].width = 25
    ws_main.column_dimensions['D'].width = 15

    ws_detail = wb.create_sheet('明细表')
    detail_headers_cn = ['原凭证号', '税种', '品目名称',
                         '税款所属时期(始)', '税款所属时期(终)', '入(退)库日期', '实缴(退)金额']
    for col_idx, header in enumerate(detail_headers_cn, 1):
        cell = ws_detail.cell(row=1, column=col_idx, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    detail_row = 2
    for rec in all_records:
        for detail in rec['details']:
            ws_detail.cell(row=detail_row, column=1, value=detail['voucher_number']).border = thin_border
            ws_detail.cell(row=detail_row, column=2, value=detail['tax_categories']).border = thin_border
            ws_detail.cell(row=detail_row, column=3, value=detail['items_name']).border = thin_border
            ws_detail.cell(row=detail_row, column=4, value=detail['start_date'] or '').border = thin_border
            ws_detail.cell(row=detail_row, column=5, value=detail['end_date'] or '').border = thin_border
            ws_detail.cell(row=detail_row, column=6, value=detail['storage_date'] or '').border = thin_border
            amount_val = 0.0
            if detail['amount']:
                try:
                    amount_val = float(detail['amount'].replace(',', ''))
                except ValueError:
                    amount_val = 0.0
            amount_cell = ws_detail.cell(row=detail_row, column=7, value=amount_val)
            amount_cell.number_format = '#,##0.00'
            amount_cell.border = thin_border
            detail_row += 1

    col_widths = [25, 25, 25, 18, 18, 18, 15]
    for i, w in enumerate(col_widths):
        ws_detail.column_dimensions[chr(65 + i)].width = w

    wb.save(output_path)
    return output_path


def generate_sql(all_records, output_path):
    """生成SQL插入语句"""
    lines = []
    lines.append('-- ============================================')
    lines.append('-- 完税证明数据 - 由app.py自动生成')
    lines.append('-- ============================================')
    lines.append('')

    lines.append('-- 主表：完税证明主表')
    lines.append('DELETE FROM invoice_main;')
    lines.append('')
    for rec in all_records:
        main = rec['main']
        inv_no = main['invoice_number'].replace("'", "''")
        amt = main['amount_tax']
        fname = rec['filename'].replace("'", "''")
        fpath = rec.get('file_path', '').replace("'", "''")
        lines.append(
            f"INSERT INTO invoice_main (file_name, file_path, invoice_number, amount_tax) "
            f"VALUES ('{fname}', '{fpath}', '{inv_no}', {amt:.2f});"
        )
    lines.append('')

    lines.append('-- 附加表：完税证明明细')
    lines.append('DELETE FROM invoice_detail;')
    lines.append('')
    for rec in all_records:
        for detail in rec['details']:
            voucher = detail['voucher_number'].replace("'", "''")
            tax_cat = detail['tax_categories'].replace("'", "''")
            items_name = detail['items_name'].replace("'", "''")
            start_date = detail['start_date'] or 'NULL'
            end_date = detail['end_date'] or 'NULL'
            storage_date = detail['storage_date'] or 'NULL'
            amount_val = 0.0
            if detail['amount']:
                try:
                    amount_val = float(detail['amount'].replace(',', ''))
                except ValueError:
                    amount_val = 0.0

            if start_date != 'NULL':
                start_date = f"'{start_date}'"
            if end_date != 'NULL':
                end_date = f"'{end_date}'"
            if storage_date != 'NULL':
                storage_date = f"'{storage_date}'"

            lines.append(
                f"INSERT INTO invoice_detail "
                f"(voucher_number, tax_categories, items_name, start_date, end_date, storage_date, amount) "
                f"VALUES ('{voucher}', '{tax_cat}', '{items_name}', "
                f"{start_date}, {end_date}, {storage_date}, {amount_val:.2f});"
            )

    with open(output_path, 'w', encoding='utf-8') as f:
        f.write('\n'.join(lines))
    return output_path


# ========== 过滤模式 ==========
MERGE_PATTERN = re.compile(r'合并')


def file_hash(filepath):
    """计算文件的MD5哈希值，用于判断文件内容是否重复"""
    h = hashlib.md5()
    with open(filepath, 'rb') as f:
        for chunk in iter(lambda: f.read(8192), b''):
            h.update(chunk)
    return h.hexdigest()


# ========== GUI应用类 ==========

class TaxExtractorApp:
    """完税证明PDF批量提取工具 - 主窗口"""

    def __init__(self):
        self.root = tk.Tk()
        self.root.title('完税证明PDF批量提取工具')
        self.root.geometry('800x600')
        self.root.minsize(700, 500)
        self.log_queue = Queue()
        self.is_running = False
        self.process_thread = None
        self._create_widgets()
        self._poll_log_queue()

    def _create_widgets(self):
        frame_src = ttk.Frame(self.root, padding=(10, 5, 10, 0))
        frame_src.pack(fill=tk.X)
        ttk.Label(frame_src, text='源文件夹：').pack(side=tk.LEFT)
        self.src_var = tk.StringVar()
        self.src_entry = ttk.Entry(frame_src, textvariable=self.src_var, width=60)
        self.src_entry.pack(side=tk.LEFT, padx=(5, 5), fill=tk.X, expand=True)
        ttk.Button(frame_src, text='浏览...', command=self._browse_source).pack(side=tk.LEFT)

        frame_opts = ttk.Frame(self.root, padding=(10, 5, 10, 0))
        frame_opts.pack(fill=tk.X)
        self.recursive_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(frame_opts, text='递归扫描子文件夹',
                        variable=self.recursive_var).pack(side=tk.LEFT, padx=(0, 20))
        self.ignore_merge_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(frame_opts, text='忽略含"合并"的文件',
                        variable=self.ignore_merge_var).pack(side=tk.LEFT)

        frame_out = ttk.Frame(self.root, padding=(10, 5, 10, 0))
        frame_out.pack(fill=tk.X)
        ttk.Label(frame_out, text='输出文件夹：').pack(side=tk.LEFT)
        desktop_path = os.path.join(os.path.expanduser('~'), 'Desktop')
        self.out_var = tk.StringVar(value=desktop_path)
        self.out_entry = ttk.Entry(frame_out, textvariable=self.out_var, width=60)
        self.out_entry.pack(side=tk.LEFT, padx=(5, 5), fill=tk.X, expand=True)
        ttk.Button(frame_out, text='浏览...', command=self._browse_output).pack(side=tk.LEFT)

        frame_name = ttk.Frame(self.root, padding=(10, 5, 10, 0))
        frame_name.pack(fill=tk.X)
        ttk.Label(frame_name, text='文件名：').pack(side=tk.LEFT)
        default_filename = datetime.now().strftime('%Y%m%d_%H%M%S_%f') + '.xlsx'
        self.filename_var = tk.StringVar(value=default_filename)
        self.filename_entry = ttk.Entry(frame_name, textvariable=self.filename_var, width=60)
        self.filename_entry.pack(side=tk.LEFT, padx=(5, 5), fill=tk.X, expand=True)

        frame_log = ttk.Frame(self.root, padding=(10, 5, 10, 5))
        frame_log.pack(fill=tk.BOTH, expand=True)
        self.log_text = tk.Text(frame_log, height=20, bg='black', fg='#00FF00',
                                font=('Consolas', 9), wrap=tk.WORD, insertbackground='green')
        scrollbar = ttk.Scrollbar(frame_log, command=self.log_text.yview)
        self.log_text.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.log_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self.log_text.configure(state=tk.DISABLED)

        frame_bottom = ttk.Frame(self.root, padding=(10, 5, 10, 10))
        frame_bottom.pack(fill=tk.X)
        self.start_btn = tk.Button(frame_bottom, text='开始提取',
                                   command=self._start_extract,
                                   bg='#4472C4', fg='white',
                                   font=('Microsoft YaHei', 10, 'bold'),
                                   relief=tk.RAISED, padx=20, pady=5)
        self.start_btn.pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(frame_bottom, text='打开输出目录',
                   command=self._open_output_dir).pack(side=tk.LEFT, padx=(0, 10))
        self.status_var = tk.StringVar(value='就绪')
        ttk.Label(frame_bottom, textvariable=self.status_var,
                  foreground='gray').pack(side=tk.LEFT, padx=(10, 0))

    def _browse_source(self):
        path = filedialog.askdirectory(title='选择PDF文件夹')
        if path:
            self.src_var.set(path)
            if self.out_var.get() == os.path.join(os.path.expanduser('~'), 'Desktop'):
                self.out_var.set(os.path.join(path, 'output'))

    def _browse_output(self):
        path = filedialog.askdirectory(title='选择输出文件夹')
        if path:
            self.out_var.set(path)

    def _open_output_dir(self):
        out_path = self.out_var.get().strip()
        if not out_path:
            messagebox.showwarning('提示', '请先设置输出文件夹')
            return
        if not os.path.isdir(out_path):
            messagebox.showwarning('提示', '输出文件夹不存在')
            return
        os.startfile(out_path)

    def _log(self, message):
        self.log_queue.put(message)

    def _poll_log_queue(self):
        # 每次最多处理50条，避免卡顿
        count = 0
        while not self.log_queue.empty() and count < 50:
            message = self.log_queue.get_nowait()
            self.log_text.configure(state=tk.NORMAL)
            self.log_text.insert(tk.END, message + '\n')
            self.log_text.see(tk.END)
            self.log_text.configure(state=tk.DISABLED)
            count += 1
        self.root.update_idletasks()
        self.root.after(30, self._poll_log_queue)

    def _start_extract(self):
        if self.is_running:
            messagebox.showinfo('提示', '正在处理中，请等待完成')
            return

        src_path = self.src_var.get().strip()
        if not src_path or not os.path.isdir(src_path):
            messagebox.showerror('错误', '请选择有效的源文件夹')
            return

        out_path = self.out_var.get().strip()
        if not out_path:
            messagebox.showerror('错误', '请设置输出文件夹')
            return

        filename = self.filename_var.get().strip()
        if not filename:
            messagebox.showerror('错误', '请输入输出文件名')
            return

        if not filename.endswith('.xlsx'):
            filename += '.xlsx'
            self.filename_var.set(filename)

        os.makedirs(out_path, exist_ok=True)

        self.log_text.configure(state=tk.NORMAL)
        self.log_text.delete('1.0', tk.END)
        self.log_text.configure(state=tk.DISABLED)

        self.is_running = True
        self.start_btn.configure(state=tk.DISABLED, bg='gray')
        self.status_var.set('处理中...')

        self.process_thread = Thread(target=self._extract_worker,
                                     args=(src_path, out_path, filename),
                                     daemon=True)
        self.process_thread.start()

    def _extract_worker(self, src_path, out_path, filename):
        start_time = time.time()

        try:
            self._log('=' * 60)
            self._log('完税证明PDF批量提取工具')
            self._log(f'源文件夹: {src_path}')
            self._log(f'输出文件夹: {out_path}')
            self._log(f'输出文件名: {filename}')
            self._log(f'递归扫描: {"是" if self.recursive_var.get() else "否"}')
            self._log(f'忽略合并文件: {"是" if self.ignore_merge_var.get() else "否"}')
            self._log('=' * 60)
            self._log('')

            if self.recursive_var.get():
                all_files = glob.glob(os.path.join(src_path, '**', '*.pdf'), recursive=True)
            else:
                all_files = glob.glob(os.path.join(src_path, '*.pdf'))

            pdf_files = []
            skipped_merge = 0
            skipped_dup = 0
            seen_hashes = set()
            for f in all_files:
                basename = os.path.basename(f)
                # 过滤"合并"文件
                if self.ignore_merge_var.get() and MERGE_PATTERN.search(basename):
                    skipped_merge += 1
                    continue
                # hash去重：内容相同的文件只保留第一个
                fhash = file_hash(f)
                if fhash in seen_hashes:
                    skipped_dup += 1
                    continue
                seen_hashes.add(fhash)
                pdf_files.append(f)

            pdf_files.sort()

            self._log(f'共发现 {len(all_files)} 个PDF文件')
            if skipped_merge > 0:
                self._log(f'  过滤合并文件: {skipped_merge} 个')
            if skipped_dup > 0:
                self._log(f'  跳过重复文件: {skipped_dup} 个')
            self._log(f'待处理文件: {len(pdf_files)} 个')
            self._log('')

            if not pdf_files:
                self._log('未找到任何PDF文件！')
                self._log('=' * 60)
                self._root_after(lambda: self._finish_extract(0, 0, 0, 0.0))
                return

            all_records = []
            success_count = 0
            fail_count = 0

            for i, pdf_path in enumerate(pdf_files):
                rel_path = os.path.relpath(pdf_path, src_path)
                self._log(f'[{i + 1}/{len(pdf_files)}] 处理: {rel_path}')

                file_start = time.time()
                try:
                    main_rec, details, fname = extract_single_pdf(pdf_path)
                    elapsed = time.time() - file_start

                    all_records.append({
                        'main': main_rec,
                        'details': details,
                        'filename': fname,
                        'file_path': os.path.relpath(pdf_path, src_path),
                    })

                    self._log(f'  -> ID: {main_rec["invoice_number"]} | 金额: {main_rec["amount_tax"]:.2f} | 明细: {len(details)}条 | {elapsed:.2f}s')
                    success_count += 1
                except Exception as e:
                    self._log(f'  ✗ 错误: {e}')
                    fail_count += 1
                self._log('')

            self._log('=' * 60)
            self._log('生成输出文件...')

            excel_path = os.path.join(out_path, filename)
            generate_excel(all_records, excel_path)
            self._log(f'Excel: {excel_path}')

            sql_filename = filename.replace('.xlsx', '.sql')
            sql_path = os.path.join(out_path, sql_filename)
            generate_sql(all_records, sql_path)
            self._log(f'SQL:   {sql_path}')

            total_elapsed = time.time() - start_time

            total_details = sum(len(r['details']) for r in all_records)
            total_amount = sum(r['main']['amount_tax'] for r in all_records)

            self._log('')
            self._log('=' * 60)
            self._log('处理完成！')
            self._log(f'  成功: {success_count} 个文件')
            self._log(f'  失败: {fail_count} 个文件')
            self._log(f'  主表记录: {success_count} 条')
            self._log(f'  明细记录: {total_details} 条')
            self._log(f'  金额合计: {total_amount:,.2f}')
            self._log(f'  总耗时: {total_elapsed:.2f}秒')
            self._log('=' * 60)

            self._root_after(lambda: self._finish_extract(
                success_count, fail_count, total_details, total_amount
            ))

        except Exception as e:
            self._log(f'处理过程中发生异常: {e}')
            import traceback
            self._log(traceback.format_exc())
            self._root_after(lambda: self._finish_extract(0, 0, 0, 0.0))

    def _root_after(self, func):
        self.root.after(0, func)

    def _finish_extract(self, success, fail, details, amount):
        self.is_running = False
        self.start_btn.configure(state=tk.NORMAL, bg='#4472C4')
        self.status_var.set(f'完成 - 成功{success}个, 失败{fail}个')

    def run(self):
        self.root.mainloop()


# ========== 入口点 ==========

if __name__ == '__main__':
    app = TaxExtractorApp()
    app.run()
