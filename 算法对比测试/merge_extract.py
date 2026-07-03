﻿# -*- coding: utf-8 -*-
"""
完税证明PDF批量提取脚本（pdfplumber + 正则表达式）
功能：
  1. 使用pdfplumber提取PDF文本和表格
  2. 通过正则表达式从文本中提取主表数据（完税证明id、完税金额）
  3. 从表格中解析明细数据（原凭证号、税种、品目名称、税款所属时期、入退库日期、实缴金额）
  4. 输出Excel（.xlsx）和SQL两种格式
"""

import os
import sys
import re
import glob
import time
import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

# ========== 配置路径 ==========
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
# 支持命令行参数指定PDF目录
if len(sys.argv) > 1:
    PDF_DIR = sys.argv[1]
else:
    # 默认使用当前目录，请通过命令行参数指定PDF目录
        PDF_DIR = os.path.join(BASE_DIR, '..', '.')
OUTPUT_DIR = os.path.join(BASE_DIR, 'output')
os.makedirs(OUTPUT_DIR, exist_ok=True)

# 去重：跳过文件名含"(1)"的副本，只处理原始文件
SKIP_PATTERN = re.compile(r'\(\d+\)\.pdf$')
# 过滤带"合并"字样的文件
MERGE_PATTERN = re.compile(r'合并')


# ========== 核心提取函数 ==========

def extract_no_from_text(text):
    """从PDF文本中提取No.字段（完税证明id）"""
    match = re.search(r'No\.\s*(\d+)', text)
    if match:
        return match.group(1)
    return None


def extract_amount_tax_from_text(text):
    """从PDF文本中提取金额合计（完税金额），格式：￥2,556.42"""
    match = re.search(r'￥([\d,]+\.?\d*)', text)
    if match:
        # 去掉逗号，转为数字字符串
        return match.group(1).replace(',', '')
    return None


def parse_date_range(date_str):
    """
    解析税款所属时期，格式如：2024-07-01至2024-07-31
    返回 (start_date, end_date)
    """
    match = re.match(r'(\d{4}-\d{2}-\d{2})至(\d{4}-\d{2}-\d{2})', date_str)
    if match:
        return match.group(1), match.group(2)
    # 尝试单个日期格式
    match = re.match(r'(\d{4}-\d{2}-\d{2})', date_str)
    if match:
        return match.group(1), match.group(1)
    return None, None


def split_merged_cell(cell_text):
    """
    pdfplumber提取的合并单元格用换行符分隔多行数据
    按换行拆分并去除空白
    """
    if not cell_text:
        return []
    lines = [line.strip() for line in cell_text.split('\n') if line.strip()]
    return lines


def extract_table_details(table_data):
    """
    从pdfplumber提取的表格原始数据中解析明细行
    表格结构（第2行开始为明细）：
      原凭证号 | (空) | 税种 | (空) | 品目名称 | (空) | 税款所属时期 | (空) | 入(退)库日期 | 实缴(退)金额
    数据行中各列可能合并，用换行符分隔多条记录
    """
    details = []
    if not table_data or len(table_data) < 3:
        return details

    # 找到数据行（跳过表头前2行：纳税人信息行 + 列标题行）
    # 在合并单元格中，多条明细合并在一个单元格里，用换行分隔
    for row in table_data:
        if not row:
            continue

        # 检查是否为数据行：第1列包含长数字（原凭证号）
        first_cell = row[0] if row[0] else ''
        # 原凭证号通常是19位数字
        if not re.match(r'.*\d{10,}', first_cell):
            continue

        # 提取各列数据（合并单元格按换行拆分）
        voucher_numbers = split_merged_cell(first_cell)
        tax_categories = split_merged_cell(row[2] if len(row) > 2 else '')
        items_names = split_merged_cell(row[4] if len(row) > 4 else '')
        date_ranges = split_merged_cell(row[6] if len(row) > 6 else '')
        storage_dates = split_merged_cell(row[8] if len(row) > 8 else '')
        amounts = split_merged_cell(row[9] if len(row) > 9 else '')

        # 取最大行数作为明细条数
        max_len = max(len(voucher_numbers), len(tax_categories), len(items_names),
                      len(date_ranges), len(storage_dates), len(amounts))

        for i in range(max_len):
            # 解析税款所属时期的起止日期
            date_str = date_ranges[i] if i < len(date_ranges) else ''
            start_date, end_date = parse_date_range(date_str)

            detail = {
                'voucher_number': voucher_numbers[i] if i < len(voucher_numbers) else '',
                'tax_categories': tax_categories[i] if i < len(tax_categories) else '',
                'items_name': items_names[i] if i < len(items_names) else '',
                'start_date': start_date,
                'end_date': end_date,
                'storage_date': storage_dates[i] if i < len(storage_dates) else '',
                'amount': amounts[i] if i < len(amounts) else '',
            }
            details.append(detail)

    return details


def extract_single_pdf(pdf_path):
    """
    提取单个PDF文件的主表和明细数据
    返回 (main_record, detail_records)
    """
    main_record = {
        'invoice_number': '',
        'amount_tax': 0.0,
    }
    detail_records = []

    filename = os.path.basename(pdf_path)

    with pdfplumber.open(pdf_path) as pdf:
        if not pdf.pages:
            return main_record, detail_records, filename

        # 通常只有一页，取第一页
        page = pdf.pages[0]
        text = page.extract_text() or ''
        tables = page.extract_tables()

        # ---- 提取主表数据 ----
        # 完税证明id：从No.字段提取
        invoice_number = extract_no_from_text(text)
        if not invoice_number:
            # 备选：从文件名中提取数字ID
            fn_match = re.search(r'(\d{10,})', filename)
            if fn_match:
                invoice_number = fn_match.group(1)
        main_record['invoice_number'] = invoice_number or ''

        # 完税金额：从文本中提取￥金额
        amount_tax_str = extract_amount_tax_from_text(text)
        main_record['amount_tax'] = float(amount_tax_str) if amount_tax_str else 0.0

        # ---- 提取明细数据 ----
        if tables:
            for table in tables:
                details = extract_table_details(table)
                detail_records.extend(details)

    return main_record, detail_records, filename


# ========== 输出函数 ==========

def generate_excel(all_records, output_path):
    """生成Excel文件，包含主表和明细两个sheet"""
    wb = Workbook()

    # ---- Sheet1: 主表 ----
    ws_main = wb.active
    ws_main.title = '主表'

    # 表头样式
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font_white = Font(bold=True, size=11, color='FFFFFF')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # 主表列头
    main_headers = ['invoice_number', 'amount_tax']
    main_headers_cn = ['完税证明ID', '完税金额']
    for col_idx, header in enumerate(main_headers_cn, 1):
        cell = ws_main.cell(row=1, column=col_idx, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    # 主表数据
    for row_idx, rec in enumerate(all_records, 2):
        main = rec['main']
        ws_main.cell(row=row_idx, column=1, value=main['invoice_number']).border = thin_border
        amount_cell = ws_main.cell(row=row_idx, column=2, value=main['amount_tax'])
        amount_cell.number_format = '#,##0.00'
        amount_cell.border = thin_border

    # 调整列宽
    ws_main.column_dimensions['A'].width = 25
    ws_main.column_dimensions['B'].width = 15

    # ---- Sheet2: 明细表 ----
    ws_detail = wb.create_sheet('明细表')

    # 明细列头
    detail_headers = ['voucher_number', 'tax_categories', 'items_name',
                      'start_date', 'end_date', 'storage_date', 'amount']
    detail_headers_cn = ['原凭证号', '税种', '品目名称',
                         '税款所属时期(始)', '税款所属时期(终)', '入(退)库日期', '实缴(退)金额']
    for col_idx, header in enumerate(detail_headers_cn, 1):
        cell = ws_detail.cell(row=1, column=col_idx, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    # 明细数据
    detail_row = 2
    for rec in all_records:
        for detail in rec['details']:
            ws_detail.cell(row=detail_row, column=1, value=detail['voucher_number']).border = thin_border
            ws_detail.cell(row=detail_row, column=2, value=detail['tax_categories']).border = thin_border
            ws_detail.cell(row=detail_row, column=3, value=detail['items_name']).border = thin_border
            ws_detail.cell(row=detail_row, column=4, value=detail['start_date'] or '').border = thin_border
            ws_detail.cell(row=detail_row, column=5, value=detail['end_date'] or '').border = thin_border
            ws_detail.cell(row=detail_row, column=6, value=detail['storage_date'] or '').border = thin_border
            amount_val = 0.0
            if detail['amount']:
                try:
                    amount_val = float(detail['amount'].replace(',', ''))
                except ValueError:
                    amount_val = 0.0
            amount_cell = ws_detail.cell(row=detail_row, column=7, value=amount_val)
            amount_cell.number_format = '#,##0.00'
            amount_cell.border = thin_border
            detail_row += 1

    # 调整列宽
    col_widths = [25, 25, 25, 18, 18, 18, 15]
    for i, w in enumerate(col_widths):
        ws_detail.column_dimensions[chr(65 + i)].width = w

    wb.save(output_path)
    return output_path


def generate_sql(all_records, output_path):
    """生成SQL插入语句，包含主表和附加表"""
    lines = []
    lines.append('-- ============================================')
    lines.append('-- 完税证明数据 - 由merge_extract.py自动生成')
    lines.append('-- ============================================')
    lines.append('')

    # ---- 主表 INSERT ----
    lines.append('-- 主表：完税证明主表')
    lines.append('DELETE FROM invoice_main;')
    lines.append('')
    for rec in all_records:
        main = rec['main']
        inv_no = main['invoice_number'].replace("'", "''")
        amt = main['amount_tax']
        lines.append(
            f"INSERT INTO invoice_main (invoice_number, amount_tax) "
            f"VALUES ('{inv_no}', {amt:.2f});"
        )
    lines.append('')

    # ---- 附加表明细 INSERT ----
    lines.append('-- 附加表：完税证明明细')
    lines.append('DELETE FROM invoice_detail;')
    lines.append('')
    for rec in all_records:
        for detail in rec['details']:
            voucher = detail['voucher_number'].replace("'", "''")
            tax_cat = detail['tax_categories'].replace("'", "''")
            items_name = detail['items_name'].replace("'", "''")
            start_date = detail['start_date'] or 'NULL'
            end_date = detail['end_date'] or 'NULL'
            storage_date = detail['storage_date'] or 'NULL'
            amount_val = 0.0
            if detail['amount']:
                try:
                    amount_val = float(detail['amount'].replace(',', ''))
                except ValueError:
                    amount_val = 0.0

            # 日期字段加引号
            if start_date != 'NULL':
                start_date = f"'{start_date}'"
            if end_date != 'NULL':
                end_date = f"'{end_date}'"
            if storage_date != 'NULL':
                storage_date = f"'{storage_date}'"

            lines.append(
                f"INSERT INTO invoice_detail "
                f"(voucher_number, tax_categories, items_name, start_date, end_date, storage_date, amount) "
                f"VALUES ('{voucher}', '{tax_cat}', '{items_name}', "
                f"{start_date}, {end_date}, {storage_date}, {amount_val:.2f});"
            )

    with open(output_path, 'w', encoding='utf-8') as f:
        f.write('\n'.join(lines))
    return output_path


# ========== 主程序 ==========

def main():
    """主函数：批量提取PDF并输出Excel和SQL"""
    print('=' * 60)
    print('完税证明PDF批量提取（pdfplumber + 正则）')
    print(f'源目录: {PDF_DIR}')
    print(f'输出目录: {OUTPUT_DIR}')
    print('=' * 60)

    # 获取所有PDF文件（递归扫描子目录，排除带括号编号的副本和"合并"文件）
    all_files = glob.glob(os.path.join(PDF_DIR, '**', '*.pdf'), recursive=True)
    pdf_files = [f for f in all_files
                 if not SKIP_PATTERN.search(os.path.basename(f))
                 and not MERGE_PATTERN.search(os.path.basename(f))]
    pdf_files.sort()

    print(f'共发现 {len(all_files)} 个PDF文件，去重后处理 {len(pdf_files)} 个')
    print()

    all_records = []
    success_count = 0
    fail_count = 0

    total_start = time.time()

    for i, pdf_path in enumerate(pdf_files):
        rel_path = os.path.relpath(pdf_path, PDF_DIR)
        print(f'[{i + 1}/{len(pdf_files)}] 处理: {rel_path}')

        start = time.time()
        try:
            main_rec, details, fname = extract_single_pdf(pdf_path)
            elapsed = time.time() - start

            all_records.append({
                'main': main_rec,
                'details': details,
                'filename': fname,
            })

            print(f'  ✓ ID: {main_rec["invoice_number"]}')
            print(f'  ✓ 金额: {main_rec["amount_tax"]:.2f}')
            print(f'  ✓ 明细: {len(details)} 条')
            print(f'  ✓ 耗时: {elapsed:.2f}秒')
            success_count += 1
        except Exception as e:
            print(f'  ✗ 错误: {e}')
            fail_count += 1
        print()

    # ---- 生成输出文件 ----
    print('=' * 60)
    print('生成输出文件...')

    # Excel输出
    excel_path = os.path.join(OUTPUT_DIR, '完税证明_提取结果.xlsx')
    generate_excel(all_records, excel_path)
    print(f'Excel: {excel_path}')

    # SQL输出
    sql_path = os.path.join(OUTPUT_DIR, '完税证明_提取结果.sql')
    generate_sql(all_records, sql_path)
    print(f'SQL:   {sql_path}')

    total_elapsed = time.time() - total_start

    # ---- 汇总统计 ----
    total_details = sum(len(r['details']) for r in all_records)
    total_amount = sum(r['main']['amount_tax'] for r in all_records)

    print()
    print('=' * 60)
    print(f'处理完成！')
    print(f'  成功: {success_count} 个文件')
    print(f'  失败: {fail_count} 个文件')
    print(f'  主表记录: {success_count} 条')
    print(f'  明细记录: {total_details} 条')
    print(f'  金额合计: {total_amount:,.2f}')
    print(f'  总耗时: {total_elapsed:.2f}秒')
    print('=' * 60)


if __name__ == '__main__':
    main()
